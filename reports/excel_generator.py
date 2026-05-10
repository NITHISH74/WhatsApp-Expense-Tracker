"""
reports/excel_generator.py
Generates multi-sheet Excel workbooks with:
  Sheet 1 — Raw expense log (decrypted)
  Sheet 2 — Category breakdown with pie chart
  Sheet 3 — Daily trends with bar chart
  Sheet 4 — Summary stats
"""

import logging
import os
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

REPORTS_DIR = Path("./reports_output")
REPORTS_DIR.mkdir(exist_ok=True)

# ─── Style constants ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")      # Dark blue
ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")     # Light blue
ACCENT_FILL = PatternFill("solid", fgColor="2E75B6")       # Medium blue
WHITE_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
BOLD_FONT = Font(bold=True, name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

CATEGORY_COLORS = [
    "4472C4", "ED7D31", "A9D18E", "FF0000", "FFC000",
    "70AD47", "255E91", "9E480E", "636363", "997300",
]


class ExcelReportGenerator:
    """
    Async-compatible Excel report generator.
    All DB reads happen in the async layer; Excel building is sync (openpyxl).
    """

    def __init__(self, db, encryption):
        self._db = db
        self._enc = encryption

    async def generate(
        self,
        phone_number: str,
        period: str = "weekly",
    ) -> str:
        """
        Generate an Excel report and return the file path.

        Args:
            phone_number: The user's phone number.
            period:       "weekly" or "monthly".

        Returns:
            Absolute path to the generated .xlsx file.
        """
        since = _period_start(period)
        expenses = await self._db.get_expenses(
            phone_number=phone_number,
            encryption=self._enc,
            since=since,
        )

        if not expenses:
            logger.info("No expenses found for %s in period %s", phone_number[:8], period)

        wb = self._build_workbook(expenses, period)
        filename = f"expense_report_{period}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = REPORTS_DIR / filename
        wb.save(file_path)
        logger.info("Report saved: %s (%d rows)", file_path, len(expenses))
        return str(file_path.resolve())

    # ─── Workbook builder ──────────────────────────────────────────────────────

    def _build_workbook(self, expenses: list[dict], period: str) -> Workbook:
        wb = Workbook()
        # Remove default sheet
        default = wb.active
        wb.remove(default)

        ws_raw = wb.create_sheet("Expense Log")
        ws_cat = wb.create_sheet("By Category")
        ws_trend = wb.create_sheet("Daily Trend")
        ws_summary = wb.create_sheet("Summary")

        self._build_raw_sheet(ws_raw, expenses)
        cat_data = self._build_category_sheet(ws_cat, expenses)
        trend_data = self._build_trend_sheet(ws_trend, expenses)
        self._build_summary_sheet(ws_summary, expenses, period)

        return wb

    # ─── Sheet 1: Raw expense log ─────────────────────────────────────────────

    def _build_raw_sheet(self, ws, expenses: list[dict]) -> None:
        ws.title = "Expense Log"
        headers = ["Date", "Time", "Category", "Description", "Amount", "Currency"]
        _write_header_row(ws, headers, row=1)

        for i, exp in enumerate(expenses, start=2):
            fill = ALT_ROW_FILL if i % 2 == 0 else None
            row_data = [
                exp["created_at"].strftime("%Y-%m-%d"),
                exp["created_at"].strftime("%H:%M"),
                exp["category"],
                exp["description"],
                exp["amount"],
                exp["currency_code"],
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if fill:
                    cell.fill = fill
                if col == 5:  # Amount column
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        # Auto-width
        _auto_width(ws, headers)
        ws.freeze_panes = "A2"

    # ─── Sheet 2: Category breakdown ─────────────────────────────────────────

    def _build_category_sheet(self, ws, expenses: list[dict]) -> dict:
        totals: dict[str, float] = defaultdict(float)
        for exp in expenses:
            totals[exp["category"]] += exp["amount"]
        sorted_cats = sorted(totals.items(), key=lambda x: x[1], reverse=True)

        headers = ["Category", "Total Amount", "% of Spend", "No. of Transactions"]
        _write_header_row(ws, headers, row=1)

        grand_total = sum(totals.values()) or 1.0
        for i, (cat, total) in enumerate(sorted_cats, start=2):
            count = sum(1 for e in expenses if e["category"] == cat)
            pct = (total / grand_total) * 100
            row_data = [cat, total, pct / 100, count]
            fill = ALT_ROW_FILL if i % 2 == 0 else None
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if fill:
                    cell.fill = fill
            ws.cell(row=i, column=2).number_format = "#,##0.00"
            ws.cell(row=i, column=3).number_format = "0.0%"

        _auto_width(ws, headers)

        # Pie chart
        if sorted_cats:
            _add_pie_chart(ws, num_rows=len(sorted_cats))

        return dict(sorted_cats)

    # ─── Sheet 3: Daily trend ─────────────────────────────────────────────────

    def _build_trend_sheet(self, ws, expenses: list[dict]) -> dict:
        daily: dict[str, float] = defaultdict(float)
        for exp in expenses:
            day_key = exp["created_at"].strftime("%Y-%m-%d")
            daily[day_key] += exp["amount"]

        sorted_days = sorted(daily.items())
        headers = ["Date", "Daily Total"]
        _write_header_row(ws, headers, row=1)

        for i, (day, total) in enumerate(sorted_days, start=2):
            ws.cell(row=i, column=1, value=day).font = BODY_FONT
            ws.cell(row=i, column=2, value=total).number_format = "#,##0.00"
            ws.cell(row=i, column=2).font = BODY_FONT

        _auto_width(ws, headers)

        if sorted_days:
            _add_bar_chart(ws, num_rows=len(sorted_days))

        return dict(sorted_days)

    # ─── Sheet 4: Summary ─────────────────────────────────────────────────────

    def _build_summary_sheet(
        self, ws, expenses: list[dict], period: str
    ) -> None:
        total = sum(e["amount"] for e in expenses)
        avg = total / len(expenses) if expenses else 0
        max_exp = max(expenses, key=lambda e: e["amount"], default=None)
        top_cat = _top_category(expenses)

        ws["A1"] = f"📊 Expense Report — {period.capitalize()}"
        ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)

        stats = [
            ("Total Expenses", len(expenses)),
            ("Total Amount", f"{total:,.2f}"),
            ("Average per Expense", f"{avg:,.2f}"),
            ("Top Category", top_cat),
            ("Largest Expense", f"{max_exp['amount']:,.2f} — {max_exp['description']}" if max_exp else "—"),
            ("Period", period.capitalize()),
            ("Report Date", datetime.utcnow().strftime("%Y-%m-%d")),
        ]

        for i, (label, value) in enumerate(stats, start=4):
            ws.cell(row=i, column=1, value=label).font = BOLD_FONT
            ws.cell(row=i, column=2, value=str(value)).font = BODY_FONT
            ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor="D6E4F0")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 40


# ─── Chart helpers ────────────────────────────────────────────────────────────

def _add_pie_chart(ws, num_rows: int) -> None:
    """Embed a pie chart below the category data."""
    pie = PieChart()
    pie.title = "Spending by Category"
    pie.style = 10

    labels = Reference(ws, min_col=1, min_row=2, max_row=num_rows + 1)
    data = Reference(ws, min_col=2, min_row=1, max_row=num_rows + 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = True
    pie.width = 18
    pie.height = 14

    ws.add_chart(pie, "F2")


def _add_bar_chart(ws, num_rows: int) -> None:
    """Embed a bar chart below the trend data."""
    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Daily Spending Trend"
    bar.y_axis.title = "Amount"
    bar.x_axis.title = "Date"
    bar.grouping = "clustered"

    data = Reference(ws, min_col=2, min_row=1, max_row=num_rows + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows + 1)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.width = 20
    bar.height = 14

    ws.add_chart(bar, "D2")


# ─── Utility helpers ──────────────────────────────────────────────────────────

def _write_header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 20


def _auto_width(ws, headers: list[str]) -> None:
    for col, header in enumerate(headers, start=1):
        letter = get_column_letter(col)
        max_len = max(
            len(str(header)),
            *(len(str(ws.cell(row=r, column=col).value or "")) for r in range(2, ws.max_row + 1)),
        )
        ws.column_dimensions[letter].width = min(max_len + 4, 50)


def _period_start(period: str) -> datetime:
    now = datetime.utcnow()
    if period == "daily":
        return now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "monthly":
        return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    else:  # weekly default
        return now - timedelta(days=7)


def _top_category(expenses: list[dict]) -> str:
    if not expenses:
        return "—"
    totals: dict[str, float] = defaultdict(float)
    for e in expenses:
        totals[e["category"]] += e["amount"]
    return max(totals, key=lambda k: totals[k])
